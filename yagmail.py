import yagmail
from openpyxl import load_workbook

#--------- DESTINATION -------------------
# Get Email's list with Excel Workbook
wb = load_workbook("emails-list.xlsx")  # Work Book
ws = wb['Plan1']  # Work Sheet
column = ws['B']  # Column
emails = [column[x].value for x in range(len(column))] # list with all e-mails
print('LISTA DE DESTINATÁRIOS: ')
print(emails)
print('\nTotal de emails a ser enviados: ' + str(len(emails)))

# Other Way - TO-list
#emails=['bla1234@emailbla.com.br', 'blabla@email.com.br']


#--------- SMTP LOGIN -------------------
# ONLY GMAIL
#yag = yagmail.SMTP('GMAIL-LOGIN', 'GMAIL-PASSWORD')

# OTHER SMTP Server
yag = yagmail.SMTP('luiseduardobr1@hotmail.com', '12345678910le', host='smtp-mail.outlook.com', port=587, smtp_starttls=True, smtp_ssl=False)

#--------- EMAIL MESSAGE AND ATTACHMENTS -------------------
# Body EMAIL
contents = [
    """Hello, 
Here is my resume.
Sincerely,
Luís Eduardo Pompeu.
""", 'curriculo_luiseduardopompeu.pdf'
]


#--------- SEND EMAIL -------------------
# Send Emails
correct=0
error=0
for endereco in emails:
    try:
        yag.send(endereco, 'Vaga Projetista - Eng. Elétrica', contents)
        print('Email enviado para: ' + str(endereco))
        correct=correct+1
    except:
        print('Problema ao enviar para: ' + str(endereco))
        error=error+1
        
print('\nFinalizado! Enviados {} emails corretamente e {} falharam'.format(correct,error))